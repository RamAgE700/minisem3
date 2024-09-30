from django.shortcuts import render, redirect
from django.views import View
from django.contrib import messages
from store.models import Category, Product, Order,Customer, Admins
from .dbset import addcat,editcat,addprodcut,eduitprodcut
from django.shortcuts import render
from store.models import Product, Order
from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404
from store.models.supplier import Supplier
from store.models.supplyreq import SupplyRequest
from store.models.SupplierProduct import SupplierProduct
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.units import mm
from reportlab.lib import colors
from datetime import timedelta
from django.contrib import messages
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from store.models.category import Category  # Import the Category model
from django.http import HttpResponseBadRequest
from django.db.models import Count
from django.db.models.functions import ExtractMonth ,ExtractYear

def index(request):
    if 'adminid' in request.session:
        unique_ids = request.session.get('adminid')
        
        # Stock management details logic
        total_products = Product.objects.count()
        total_orders = Order.objects.count()
        total_users = User.objects.count()
        total_profit = sum(order.price for order in Order.objects.all())
        
        # Get product names and stock counts
        products_stock = Product.objects.values('name', 'quantity')

        context = {
            'total_products': total_products,
            'total_orders': total_orders,
            'total_users': total_users,
            'total_profit': total_profit,
            'products_stock': products_stock,  # Add this line
        }
       
        return render(request, 'admins/homes.html', context)
    else:
        return redirect('adlogin')
def logouts(request):
      if 'adminid' in request.session:
             del request.session['adminid']
             del request.session['bkadmin']
             return redirect('adlogin')
      return render(request,'admins/login.html')


def addcats(request):
    if request.method == "POST":
        category_name = request.POST.get('cate')
        if Category.objects.filter(name=category_name).exists():
            messages.warning(request, "Category already exists.")
        else:
            Category.objects.create(name=category_name)
            return redirect('addcats')  # Redirect to the same page to clear the form and display the success message

    return render(request, 'admins/cpaddcat.html')  # Redirect to the same page to clear the form and display the success message

    return render(request, 'admins/cpaddcat.html')
from django.contrib.auth.decorators import login_required


def addproduct(request):
    cathome = Category.objects.all()
    if request.method == "POST":
        sname = request.POST.get('sname')
        sp = request.POST.get('sp')
        sprce = request.POST.get('sprce')
        sdic = request.POST.get('sdic')
        prdimg = request.FILES.get('prdimg')
        spqty = request.POST.get('spqty')
        
        # Check if a product with the same name and category already exists
        if Product.objects.filter(name=sname, category_id=sp).exists():
            messages.warning(request, "Product already exists.")
        else:
            # Add the new product
            new_product = Product(name=sname, category_id=sp, price=sprce, description=sdic, image=prdimg, quantity=spqty )
            new_product.save()
            messages.success(request, "Product added successfully.")
            return redirect('viewproduct')    # Redirect to a page showing the list of products

    return render(request, 'admins/addprodcut.html', {'c': cathome})

def editproduct(request):
       cats=Category.objects.values()
       delit= request.GET.get('keys', None)
       cathome=Product.objects.filter(id=delit)
       if request.method=="POST":
            if eduitprodcut(request,delit)==1:
               messages.success(request, "info Updated.")
                  
            else:
                  messages.success(request, "Something error.")     
       return render(request,'admins/editprodcut.html',{ 'pd': cathome.all(),'c':cats})
def editcats(request):
      delit= request.GET.get('edit', None)
      cathome=Category.objects.filter(id=delit)
      if request.method=="POST":
                  if editcat(request,delit)==1 :
                        messages.success(request, "Something error.")     
      return render(request,'admins/editcat.html',{'bk':cathome.all()})
 
def viewcat(request):
       cathome=Category.objects.values()
       delit= request.GET.get('del', None)
       if delit is not None:
                        instance = Category.objects.get(id=delit)
                        instance.delete()
                        
       return render(request,'admins/catlist.html',{'bk':cathome.all()})
def viewprd(request):
       cathome=Product.objects.values()
       delit= request.GET.get('del', None)
       if delit is not None:
                        instance = Product.objects.get(id=delit)
                        instance.delete()
                        
       return render(request,'admins/prodcutlist.html',{'z':cathome.all()})
def viewporder(request):
    cathome = Order.objects.select_related('customer')
    prd = Product.objects.values()
    cust = Customer.objects.values()
    delit = request.GET.get('del', None)
    status_update = request.GET.get('status_update', None)
    new_status = request.GET.get('status', None)

    if delit is not None:
        instance = Order.objects.get(id=delit)
        instance.delete()
    if status_update is not None and new_status is not None:
        instance = Order.objects.get(id=status_update)
        instance.status = new_status
        instance.save()

    return render(request, 'admins/order.html', {'z': cathome.all(), 'prd': prd.all(), 'cust': cust.all()})




from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.db.models import Sum, F


def game_sales_report(request):
    # Get the month from the query parameters, default to the current month and year
    month = request.GET.get('month', datetime.now().strftime('%Y-%m'))
    
    # Split the month and year for query
    year, month = map(int, month.split('-'))

    # Calculate the start and end dates for the selected month
    start_date = datetime(year, month, 1)
    if month < 12:
        end_date = datetime(year, month + 1, 1)
    else:
        end_date = datetime(year + 1, 1, 1)

    # Filter orders within the selected month
    orders = Order.objects.filter(date__range=[start_date, end_date])
    
    # Calculate total sales for the month
    total_sales = orders.aggregate(total=Sum('price'))['total'] or 0

    # Calculate sales per game for the month
    product_sales = orders.values('product__name').annotate(
        total_quantity=Sum('quantity'),
        total_amount=Sum(F('quantity') * F('price'))
    ).order_by('-total_amount')

    # Calculate the unit price for each game
    for item in product_sales:
        item['unit_price'] = item['total_amount'] / item['total_quantity'] if item['total_quantity'] else 0

    # Export to Excel if requested
    if 'export' in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.title = f"Game Sales Report {month}-{year}".replace("/", "-")

        # Write the headers
        headers = ['Game', 'Unit Price', 'Quantity Sold', 'Total Sales Amount']
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

        # Write data to the Excel sheet
        for row_num, item in enumerate(product_sales, 2):
            ws[f"A{row_num}"] = item['product__name']
            ws[f"B{row_num}"] = round(item['unit_price'], 2)
            ws[f"C{row_num}"] = item['total_quantity']
            ws[f"D{row_num}"] = round(item['total_amount'], 2)

        # Add a row for total sales
        ws[f"A{row_num + 1}"] = 'Total Sales'
        ws[f"D{row_num + 1}"] = total_sales

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Game_Sales_Report_{month}_{year}.xlsx'
        wb.save(response)
        return response

    context = {
        'total_sales': total_sales,
        'month': f"{year}-{month:02d}",
        'product_sales': product_sales,
    }
    return render(request, 'admins/game_sales_report.html', context)

from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, F

import matplotlib.pyplot as plt
import io
import urllib, base64

def visualizations(request):
    month = request.GET.get('month', datetime.now().strftime('%Y-%m'))
    year, month = map(int, month.split('-'))

    # Calculate the start and end dates for the selected month
    start_date = datetime(year, month, 1)
    if month < 12:
        end_date = datetime(year, month + 1, 1)
    else:
        end_date = datetime(year + 1, 1, 1)

    # Filter orders within the selected month
    orders = Order.objects.filter(date__range=[start_date, end_date])
    
    if not orders.exists():
        print("No orders found for this month.")
    
    # Calculate sales per game for the month
    product_sales = orders.values('product__name').annotate(
        total_quantity=Sum('quantity'),
        total_amount=Sum(F('quantity') * F('price'))
    ).order_by('-total_amount')

    if not product_sales.exists():
        print("No product sales data found.")

    # Generate Bar Chart for Sales by Game
    game_names = [item['product__name'] for item in product_sales]
    total_amounts = [item['total_amount'] for item in product_sales]

    # Ensure there are items to plot
    if not game_names:
        print("No game names found for chart.")
        game_names = ["No Data"]
        total_amounts = [0]

    plt.figure(figsize=(10, 5))
    plt.bar(game_names, total_amounts, color='skyblue')
    plt.xlabel('Games')
    plt.ylabel('Total Sales ($)')
    plt.title(f'Sales by Game for {year}-{month:02d}')
    plt.xticks(rotation=45, ha='right')

    # Save plot to image
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    string = base64.b64encode(buf.read())
    uri = urllib.parse.quote(string)

    context = {
        'month': f"{year}-{month:02d}",
        'sales_chart': uri,  # Pass the chart URI to the template
    }
    
    # Debug: Print context data
    print("Context Data:", context)
    
    return render(request, 'admins/visualizations.html', context)



 
def adminlogin(request):
     context=''
     
     if request.method=="POST":
        ausername=request.POST.get('username')
        apassword=request.POST.get('password')
        user=Admins.objects.filter(username=ausername, password=apassword ).values()
        if user.count() == 1:
               
                current_user = user.get()['id']
                request.session['bkadmin']= current_user
                request.session['adminid']= user.get()['username']
                return redirect('index')
        else:
                context ='Invalid email or password'
                return render(request, 'admins/login.html', {'context': context, 'error': 'Invalid email or password'})
              
     return render(request,'admins/login.html', {'context': context})

def viewBill(request):
              customer_id = request.session.get('customer')
              sustomers=Customer.objects.filter(id=customer_id)
              orders = Order.objects.filter(customer=customer_id).order_by("-date").order_by("-id")
              
              return render(request,'bill.html',{"orders":orders, 'ck': sustomers})

class InStockPageView(View):
    template_name = 'admins/in_stock.html'

    def get(self, request):
        cats = Category.objects.all()
        supplier_products = SupplierProduct.objects.all()
        context = {
            'in_stock_products': supplier_products,
            'c': cats
        }
        return render(request, self.template_name, context)

    def post(self, request):
        product_id = request.POST.get('product_id')  # Retrieve product_id from the POST data
        if product_id is None:
            return HttpResponseBadRequest("Product ID is missing in the request.")
        
        supplier_product = get_object_or_404(SupplierProduct, id=product_id)
        
        total_price = float(request.POST.get('price'))
        category_id = request.POST.get('sp')
        category = Category.objects.get(id=category_id)
        
        new_product = Product(
            name=supplier_product.product_name,
            price=total_price,
            category=category,
            description=supplier_product.description,
            image=supplier_product.image,
            quantity=supplier_product.number_of_items,
            Supplier=supplier_product.Supplier
        )
        
        new_product.save()
        
        return redirect('in_stock_page')  # Redirect to avoid form resubmission


def get_supplier_product_pdf(request, product_id):
    # Retrieve the supplier product by ID
    supplier_product = get_object_or_404(SupplierProduct, id=product_id)
    
    # Create the HTTP response object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Order_{product_id}_Bill.pdf"'

    
    # Create a SimpleDocTemplate for the PDF
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []

    styles = getSampleStyleSheet()
    
    # Invoice header
    elements.append(Spacer(1, 20))
    header_data = [["RIOT STORE", ""], ["RIOT STORE Building ", ""]]
    header_table = Table(header_data, colWidths=[150*mm, 70*mm], hAlign='LEFT')
    elements.append(header_table)
    
    elements.append(Spacer(1, 40))
    
    # Supplier Name
    supplier_name_data = [["Supplier Name:", supplier_product.Supplier.name]]
    supplier_name_table = Table(supplier_name_data, colWidths=[90*mm, 100*mm])
    supplier_name_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))
    elements.append(supplier_name_table)

    elements.append(Spacer(1, 20))

    # Supplier Product Information
    product_info_data = [
        ["Order ID:", supplier_product.id],
        ["Product Name:", supplier_product.product_name],
        ["Price:", f"{supplier_product.price:.2f}"],
        ["Supply Charge:", f"{supplier_product.supply_charge:.2f}"],
        ["Number of Items:", supplier_product.number_of_items],
        ["Date of Supply:", supplier_product.date_of_supply.strftime('%d/%m/%Y')],
        ["Status:", supplier_product.status],
        ["Supplier:", supplier_product.Supplier.name],  # Add Supplier field
         # Add Supplier address field
    ]
    product_info_table = Table(product_info_data, colWidths=[90*mm, 100*mm])
    product_info_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('ALIGN', (1, 0), (-1, -1), 'LEFT'),
    ]))
    elements.append(product_info_table)
    
    elements.append(Spacer(1, 20))

    # Total amount or other summary if necessary
    total_data = [["", "", "TOTAL:", f"{supplier_product.price * supplier_product.number_of_items + supplier_product.supply_charge:.2f}"]]
    total_table = Table(total_data, colWidths=[90*mm, 30*mm, 30*mm, 30*mm])
    total_table.setStyle(TableStyle([
        ('BACKGROUND', (2, 0), (-1, 0), colors.lightgrey),
        ('ALIGN', (2, 0), (-1, 0), 'CENTER'),
    ]))
    elements.append(total_table)
    
    elements.append(Spacer(1, 60))
    
    # Signature
    signature_data = [["Issued by, signature:", "", "Cool breeze shop owner"]]
    signature_table = Table(signature_data, colWidths=[50*mm, 90*mm, 50*mm])
    elements.append(signature_table)
    
    # Build the PDF
    doc.build(elements)
    return response


from django.http import HttpResponseNotAllowed


def delete_product(request, product_id):
    if request.method == 'POST':
        # Retrieve the SupplierProduct object based on the product_id
        supplier_product = get_object_or_404(SupplierProduct, pk=product_id)
        # Delete the product
        supplier_product.delete()
        # Redirect to a relevant URL after deletion
        return redirect('in_stock_page')  # Adjust the URL name as per your URL configuration
    else:
        return HttpResponseNotAllowed(['POST'])  # Return a 405 Method Not Allowed for GET requests


def send_bill(request):
    # Your logic for sending the bill
    return HttpResponse("Bill sent successfully!") 





# views/admincnt.py

from django.db.models import Sum

def view_bill(request):
    # Retrieve products that are in stock from both Product and SupplierProduct models
    admin_products = Product.objects.filter(in_stock=True)
    supplier_products = SupplierProduct.objects.all()

    # Combine both sets of products
    in_stock_products = list(admin_products) + list(supplier_products)

    # Prepare context data
    context = {
        'in_stock_products': in_stock_products
    }

    # Render the template with the context data
    return render(request, 'admins/in_stock.html', context)




